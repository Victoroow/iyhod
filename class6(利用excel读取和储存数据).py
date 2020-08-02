import openpyxl
# python操作excel的三大对象
# 1.加载工作薄（excel）
# 先把excel放到 python文件的同级目录
# wb = openpyxl.load_workbook('test_case_api.xlsxtest_case_api.xlsx')
# print(wb)

# # 2.操作sheet表单
# sheet = wb['register']
# print(wb)

# # 3.操作单元格
# # row=行 column= 列
# cell = sheet.cell(row=2,column=5)

# # 取出的是单元格
# print(cell)
# #  取出的单元格的数据，要加上.value
# sheet = wb['register']
# print(sheet)

# # 写入数据
# # value赋值
# cell.value = '测试用例编号'
# print(cell.value)

# # 把写入的数据进行保存
# wb.save('test_case_api.xlsx')

# 自动化读取excel
import requests
import openpyxl

# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# for i in range(2,8,1):  # 取值取左不取右，左闭右开
#     url = sheet.cell(row= i,column= 5).value  #取出Url
#     data = sheet.cell(row= i,column= 6).value #取出请求体
#     expected = sheet.cell(row= i,column= 7).value #取出预期结果
#     print(url,data,expected)

# def logon_func(url,data):
#     header_login = {'X-Lemonban-Media-Type': 'lemonban.v2',
#                        'Content-Type': 'application/json'}
#     res1 = requests.post(url=url, json=data, headers=header_login)
#     print(res1.json())
# res = logon_func(url,data)
# print(res)

# max_row 最大的行数
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# max_row = sheet.max_row  #取出sheet最大的行数
# # print(max_row)
# for i in range(2,max_row+1,1):
#     dict1 = dict(
#     url = sheet.cell(row= i,column= 5).value  #取出Url
#     data = sheet.cell(row= i,column= 6).value #取出请求体
#     expected = sheet.cell(row= i,column= 7).value #取出预期结果
#      )
#     case_list.append(dict1) # dict1里面是一条 一条的测试用例   ---》装到列表里面，这个列表就存放了所有的测试用例
# print(case_list)

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 取出sheet最大的行数
    # print(max_row)
    case_list = []
    for i in range(2, max_row+1, 1):
        dict1 = dict(
        url=sheet.cell(row=i, column=5).value, # 取出Url
        data = sheet.cell(row=i, column=6).value, # 取出请求体
        expected = sheet.cell(row=i, column=7).value  # 取出预期结果
        )
        case_list.append(dict1)  # dict1里面是一条 一条的测试用例   ---》装到列表里面，这个列表就存放了所有的测试用例
    print(case_list)
read_data('test_case_api.xlsx','register')